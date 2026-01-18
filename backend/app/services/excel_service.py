"""Excel 读写服务模块"""
import os
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models.schemas import AccountModel, RegistrationStatus
from app.services.crypto_service import CryptoService
from app.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelService:
    """Excel 读写服务类"""

    def __init__(self, file_path: str, crypto_service: Optional[CryptoService] = None):
        """
        初始化 Excel 服务

        Args:
            file_path: Excel 文件路径
            crypto_service: 加密服务实例
        """
        self.file_path = file_path
        self.crypto = crypto_service or CryptoService()
        self.workbook = None
        self.sheet = None

        # 列名映射 (中文 -> 英文字段)
        self.column_map = {
            'ID': 'id',
            '邮箱': 'email',
            '密码': 'password',
            '备用邮箱': 'recovery_email',
            '代理地址': 'proxy_host',
            '代理端口': 'proxy_port',
            '代理用户名': 'proxy_username',
            '代理密码': 'proxy_password',
            '状态': 'status',
            'GPT账号ID': 'gpt_account_id',
            'GPT邮箱': 'gpt_email',
            '错误信息': 'error_message',
            '重试次��': 'retry_count',
            '最后尝试时间': 'last_attempt',
            '创建时间': 'created_at',
            '更新时间': 'updated_at'
        }

        # 反向映射 (英文 -> 中文)
        self.reverse_column_map = {v: k for k, v in self.column_map.items()}

    async def load_accounts(self) -> List[AccountModel]:
        """
        从 Excel 加载账号列表

        Returns:
            账号模型列表
        """
        try:
            if not os.path.exists(self.file_path):
                logger.warning(f"Excel 文件不存在: {self.file_path}")
                return []

            self.workbook = load_workbook(self.file_path)
            self.sheet = self.workbook.active

            accounts = []

            # 跳过表头,从第2行开始
            for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
                # 邮箱列为空则跳过 (索引1)
                if not row or len(row) < 2 or not row[1]:
                    continue

                try:
                    account = AccountModel(
                        id=row[0] if len(row) > 0 else None,
                        email=row[1],
                        password=self.crypto.decrypt(row[2]) if len(row) > 2 and row[2] else '',
                        recovery_email=row[3] if len(row) > 3 else None,
                        proxy_host=row[4] if len(row) > 4 else None,
                        proxy_port=row[5] if len(row) > 5 else None,
                        proxy_username=row[6] if len(row) > 6 else None,
                        proxy_password=self.crypto.decrypt(row[7]) if len(row) > 7 and row[7] else None,
                        status=RegistrationStatus(row[8]) if len(row) > 8 and row[8] else RegistrationStatus.PENDING,
                        gpt_account_id=row[9] if len(row) > 9 else None,
                        gpt_email=row[10] if len(row) > 10 else None,
                        error_message=row[11] if len(row) > 11 else None,
                        retry_count=row[12] if len(row) > 12 else 0,
                        last_attempt=row[13] if len(row) > 13 else None,
                        created_at=row[14] if len(row) > 14 else datetime.now(),
                        updated_at=row[15] if len(row) > 15 else datetime.now()
                    )
                    accounts.append(account)
                except Exception as e:
                    logger.error(f"解析第 {row_idx} 行失败: {str(e)}, 行数据: {row}")
                    continue

            logger.info(f"从 Excel 加载了 {len(accounts)} 个账号")
            return accounts

        except Exception as e:
            logger.error(f"加载 Excel 失败: {str(e)}")
            raise

    async def save_accounts(self, accounts: List[AccountModel]):
        """
        保存账号列表到 Excel

        Args:
            accounts: 账号模型列表
        """
        try:
            # 创建或打开工作簿
            if os.path.exists(self.file_path):
                self.workbook = load_workbook(self.file_path)
            else:
                self.workbook = Workbook()

            self.sheet = self.workbook.active

            # 写入表头
            headers = list(self.column_map.keys())
            for col_idx, header in enumerate(headers, 1):
                cell = self.sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 写入数据
            for row_idx, account in enumerate(accounts, 2):
                status_value = (
                    account.status.value
                    if hasattr(account.status, "value")
                    else str(account.status)
                )
                self.sheet.cell(row=row_idx, column=1, value=account.id)
                self.sheet.cell(row=row_idx, column=2, value=account.email)
                self.sheet.cell(row=row_idx, column=3, value=self.crypto.encrypt(account.password))
                self.sheet.cell(row=row_idx, column=4, value=account.recovery_email)
                self.sheet.cell(row=row_idx, column=5, value=account.proxy_host)
                self.sheet.cell(row=row_idx, column=6, value=account.proxy_port)
                self.sheet.cell(row=row_idx, column=7, value=account.proxy_username)
                self.sheet.cell(row=row_idx, column=8,
                                value=self.crypto.encrypt(account.proxy_password) if account.proxy_password else None)
                self.sheet.cell(row=row_idx, column=9, value=status_value)
                self.sheet.cell(row=row_idx, column=10, value=account.gpt_account_id)
                self.sheet.cell(row=row_idx, column=11, value=account.gpt_email)
                self.sheet.cell(row=row_idx, column=12, value=account.error_message)
                self.sheet.cell(row=row_idx, column=13, value=account.retry_count)
                self.sheet.cell(row=row_idx, column=14, value=account.last_attempt)
                self.sheet.cell(row=row_idx, column=15, value=account.created_at)
                self.sheet.cell(row=row_idx, column=16, value=account.updated_at)

                # 根据状态设置单元格背景色
                status_cell = self.sheet.cell(row=row_idx, column=9)
                if status_value == RegistrationStatus.SUCCESS.value:
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif status_value == RegistrationStatus.FAILED.value:
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif status_value == RegistrationStatus.IN_PROGRESS.value:
                    status_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                elif status_value == RegistrationStatus.PENDING.value:
                    status_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # 自动调整列宽
            for column in self.sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                self.sheet.column_dimensions[column_letter].width = adjusted_width

            # 确保目录存在
            os.makedirs(os.path.dirname(self.file_path) or '.', exist_ok=True)

            # 保存文件
            self.workbook.save(self.file_path)
            logger.info(f"保存 {len(accounts)} 个账号到 Excel: {self.file_path}")

        except Exception as e:
            logger.error(f"保存 Excel 失败: {str(e)}")
            raise

    async def update_account_status(
            self,
            email: str,
            status: RegistrationStatus,
            error_message: Optional[str] = None,
            gpt_account_id: Optional[str] = None,
            gpt_email: Optional[str] = None
    ):
        """
        更新单个账号状态

        Args:
            email: 账号邮箱
            status: 新状态
            error_message: 错误信息
            gpt_account_id: GPT账号ID
            gpt_email: GPT邮箱
        """
        accounts = await self.load_accounts()

        for account in accounts:
            if account.email == email:
                account.status = status
                account.error_message = error_message
                account.gpt_account_id = gpt_account_id
                account.gpt_email = gpt_email
                account.updated_at = datetime.now()
                logger.info(f"更新账号状态: {email} -> {status.value}")
                break

        await self.save_accounts(accounts)
